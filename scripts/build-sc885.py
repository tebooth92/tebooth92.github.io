#!/usr/bin/env python3
"""Extend pay-data.json with SC-885 (2025) Health & Allied Services / Managers & Admin rates."""
import openpyxl, json, sys
from datetime import date, timedelta

XLSX = '/sessions/affectionate-friendly-turing/mnt/uploads/SC-885_2025-Health-Allied-Services-Managers-Admin-a024c09f.xlsx'
PAYDATA = '/sessions/affectionate-friendly-turing/repo/pay-data.json'
HRS_WK, HRS_DAY = 38, 7.6
NEW_START = date(2025, 12, 16)

def fmt(x): return ('%.5f' % round(x, 5)).rstrip('0').rstrip('.')
def hourly_wk(w): return fmt(w / HRS_WK)
def hourly_day(d): return fmt(d / HRS_DAY)
def iso(d): return d.isoformat()

ALLOW_DATES = [(date(2026,1,1), 3), (date(2027,1,1), 6), (date(2027,12,31), 7)]
ALLOW_MAP = {
    'Experience Payment after Yr 1': (6, 'wk'),
    'Experience Payment after Yr 2': (7, 'wk'),
    'Experience Payment after Yr 3': (8, 'wk'),
    'Experience Payment after Yr 4': (9, 'wk'),
    'Uniform Allowance': (20, 'wk'),
    'Laundry Allowance': (24, 'wk'),
    'TOW Motor Allowance': (31, 'day'),
    'Cert Pathology Technician': (29, 'wk'),
    'Security Cert Allowance': (37, 'wk'),
    'Cert III Non-Emergency PT': (34, 'wk'),
    'Cert IV in Health Care': (35, 'wk'),
    'Diploma of Paramedical Sc': (36, 'wk'),
}

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Wages']; wa = wb['Allowances']

allow_vals = {}
for name, (row, basis) in ALLOW_MAP.items():
    allow_vals[name] = {}
    for d, col in ALLOW_DATES:
        v = wa.cell(row, col).value
        if isinstance(v, (int, float)):
            allow_vals[name][d] = hourly_wk(v) if basis == 'wk' else hourly_day(v)

WAGE_COLS = [(date(2025,12,16),4),(date(2026,1,1),5),(date(2026,7,1),8),(date(2026,12,16),9),(date(2027,12,16),12)]
wage_events = {}; translated = set()
for r in range(2, ws.max_row+1):
    pc = ws.cell(r,3).value
    if not pc: continue
    pc = str(pc).strip()
    if ws.cell(r,7).value in (None, ''): translated.add(pc)
    ev = {}
    for d, col in WAGE_COLS:
        v = ws.cell(r, col).value
        if isinstance(v, (int, float)): ev[d] = v
    wage_events[pc] = ev

data = json.load(open(PAYDATA)); codes = data['codes']
report = []
for pc, ev in wage_events.items():
    if pc not in codes: print('MISSING', pc); continue
    cd = codes[pc]; ranges = cd['dateRanges']; last = ranges[-1]
    assert last['endDate'] == '9999-12-31', f'{pc}: last not open'
    assert last['startDate'] < iso(NEW_START), f'{pc}: starts {last["startDate"]}'
    template = [dict(c) for c in last['components']]
    names_order = [c['name'] for c in template]
    cur = {c['name']: c['amount'] for c in template}
    last['endDate'] = iso(NEW_START - timedelta(days=1))
    if pc in translated:
        base = hourly_wk(ev[NEW_START]); cur['Base Pay Rate'] = base
        if 'Casual Rate' in cur: cur['Casual Rate'] = base
        ranges.append({'startDate': iso(NEW_START), 'endDate': '9999-12-31',
                       'components': [{'name': n, 'amount': cur[n]} for n in names_order]})
        report.append((pc, 1, 'translated')); continue
    applicable_allow = [d for d,_ in ALLOW_DATES if any(n in cur and d in allow_vals.get(n,{}) for n in ALLOW_MAP)]
    all_dates = sorted(set(list(ev.keys()) + applicable_allow))
    all_dates = [d for d in all_dates if d >= NEW_START]
    for i, d in enumerate(all_dates):
        if d in ev:
            base = hourly_wk(ev[d]); cur['Base Pay Rate'] = base
            if 'Casual Rate' in cur: cur['Casual Rate'] = base
        for n in ALLOW_MAP:
            if n in cur and d in allow_vals.get(n, {}): cur[n] = allow_vals[n][d]
        end = '9999-12-31' if i == len(all_dates)-1 else iso(all_dates[i+1] - timedelta(days=1))
        ranges.append({'startDate': iso(d), 'endDate': end,
                       'components': [{'name': n, 'amount': cur[n]} for n in names_order]})
    report.append((pc, len(all_dates), 'standard'))

data['metadata']['version'] = '2.3.0'
data['metadata']['generatedAt'] = '2026-06-26'
ags = data['metadata']['agreements']
SC885 = 'Health & Allied Services / Managers & Administrative Workers (SC-885)'
if SC885 not in ags: ags.append(SC885)
data['metadata'].setdefault('sources', {})['SC-885'] = {
    'circular': 'SC-885 - 2025',
    'agreement': 'Health and Allied Services, Managers and Administrative Workers (Victorian Public Sector) (Single Interest Employers) Enterprise Agreement 2025',
    'hoursPerWeek': 38,
    'note': ('Weekly rates -> hourly via /38; per-day allowances (TOW Motor) -> hourly via /7.6. '
             'Casual Rate mirrors Base Pay Rate (no loading). Wage increases: 16 Dec 2025, '
             'off-cycle 1 Jan 2026 (PS25) and 1 Jul 2026 (Support Services Level 1 codes), '
             '16 Dec 2026, 16 Dec 2027. Allowance increases on a separate timeline (1 Jan 2026, '
             '1 Jan 2027, 31 Dec 2027) interleaved into dateRanges. Translated/superseded codes '
             '(RH3-5 -> RH2, RG7-9 -> RG6) kept with their 16 Dec 2025 rate only.'),
    'sourceFile': 'SC-885_2025-Health-Allied-Services-Managers-Admin.xlsx',
    'effectiveFrom': '2025-12-16',
}

mode = sys.argv[1] if len(sys.argv) > 1 else 'validate'
if mode == 'apply':
    json.dump(data, open(PAYDATA, 'w'), separators=(',', ':'), ensure_ascii=False)
    print('WROTE', PAYDATA)
print(f'\nProcessed {len(report)} codes; translated={len(translated)} {sorted(translated)}')
print('total new ranges:', sum(r[1] for r in report))
